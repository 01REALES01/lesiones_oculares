"""
API principal: plataforma de análisis de retinografías.

- Comparación de modelos de retinopatía diabética sobre la misma imagen.
- Trazabilidad de inferencias e historial básico.
- Postprocesamiento: etiquetas, probabilidades y tiempos de inferencia.
- Uso orientado a apoyo clínico/educativo (no diagnóstico).
"""

import asyncio
import time
import uuid
from typing import Any, List, Optional
from datetime import timedelta, datetime, timezone

import numpy as np
import cv2
from fastapi import FastAPI, File, Query, UploadFile, Depends, HTTPException, status, Request
from fastapi.responses import JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.staticfiles import StaticFiles

from backend.roble_db import roble_insert
from backend.roble_db import ensure_user_exists
from backend.roble_db import save_analysis_to_roble
from backend.roble_db import roble_delete_record
from backend.roble_db import list_user_analyses_from_roble
from backend.roble_db import roble_delete_batch
from backend.roble_db import roble_delete_all_user_history
from backend.roble_db import get_user_stats_from_roble
from backend.roble_db import get_analysis_from_roble, get_batch_from_roble

try:
    import anthropic as anthropic_sdk
except ImportError:
    anthropic_sdk = None

import openpyxl
from io import BytesIO
from fastapi.responses import StreamingResponse
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import styles
from datetime import datetime

from backend.config import settings
from backend.preprocessing.fundus import crop_img, preprocess_fundus
from backend.models.segmentation_vnet import segment_optic_disc
from backend.models.glaucoma_classifier import predict_glaucoma
from backend.models.lesion_detector import detect_hemorrhages
from backend.postprocessing.report import build_report, graph_data_for_frontend
from backend.store import (
    save_image_to_disk2,
    save_inference,
    save_inferences_batch,
    get_inference,
    list_inferences,
    get_global_stats,
    save_image_to_disk,
    clear_images_dir,
    get_batch,
    clear_history,
    delete_batch,
    IMAGES_DIR,
)
from backend.ml_manager import ml_manager
from backend.agents import brain_agent, AgentAnalysisResponse
from contextlib import asynccontextmanager
# Auth imports
from backend.auth import (
    Token, User, get_current_user, create_access_token, 
    get_user, verify_password, oauth2_scheme, ACCESS_TOKEN_EXPIRE_MINUTES
)


import httpx

MODEL_RUNTIME_METADATA = {}

# Disclaimer fijo (apoyo/tamizaje, no diagnóstico)
DISCLAIMER = (
    "Este sistema es de apoyo clínico y educativo. No constituye diagnóstico médico. "
    "Los resultados deben ser interpretados por un profesional de la salud. "
    "No usar como único criterio para decisiones clínicas."
)

RD_MODEL_SPECS = {
    "densenet169": {
        "model_key": "lesiones_densenet",
        "model_type": "densenet169",
        "label": "DenseNet169",
    },
    "resnet50": {
        "model_key": "lesiones_resnet",
        "model_type": "resnet50",
        "label": "ResNet50",
    },
    "xception": {
        "model_key": "lesiones_xception",
        "model_type": "xception",
        "label": "Xception",
    },
    "efficientnet": {
        "model_key": "lesiones_efficientnet",
        "model_type": "efficientnet",
        "label": "EfficientNet",
    },
    "mobilenetv3": {
        "model_key": "lesiones_mobilenet",
        "model_type": "mobilenetv3",
        "label": "MobileNetV3",
    },
}


def _load_first_available_model(model_key: str, candidates: List[dict]) -> None:
    last_error = None
    for candidate in candidates:
        filename = candidate["filename"]
        try:
            ml_manager.load_model(model_key, filename)
            MODEL_RUNTIME_METADATA[model_key] = {
                "filename": filename,
                "model_type": candidate.get("model_type"),
                "label": candidate.get("label", filename),
            }
            print(f"Model '{model_key}' loaded from '{filename}'.")
            return
        except Exception as exc:
            last_error = exc
    print(f"Warning: Could not load model '{model_key}' from candidates {[item['filename'] for item in candidates]}: {last_error}")

@asynccontextmanager
async def lifespan(app: FastAPI):
    # Load available DR models. Fallback per model is handled at inference time.
    _load_first_available_model("lesiones_densenet", [
        {"filename": "densenet_169_aptos_fine.h5", "model_type": "densenet169", "label": "DenseNet169"},
    ])
    _load_first_available_model("lesiones_resnet", [
        {"filename": "resnet50_model_fine.keras", "model_type": "resnet50", "label": "ResNet50 Fine"},
        {"filename": "resnet50_model_fine.h5", "model_type": "resnet50", "label": "ResNet50"},
    ])
    _load_first_available_model("lesiones_xception", [
        {"filename": "xception_aptos_fine2.h5", "model_type": "xception", "label": "Xception"},
        {"filename": "xception_model.keras", "model_type": "xception", "label": "Xception"},
    ])
    _load_first_available_model("lesiones_efficientnet", [
        {"filename": "efficientnetB0_weights.h5", "model_type": "efficientnet", "label": "EfficientNetB0"},
        {"filename": "efficientnet_model.h5", "model_type": "efficientnet", "label": "EfficientNet"},
        {"filename": "efficientnet_weights.h5", "model_type": "efficientnet", "label": "EfficientNet"},
        {"filename": "efficientnetB0.keras", "model_type": "efficientnet", "label": "EfficientNetB0"},
    ])
    _load_first_available_model("lesiones_mobilenet", [
        {"filename": "mobilenetv3_model_fine.keras", "model_type": "mobilenetv3", "label": "MobileNetV3 Fine"},
        {"filename": "mobilenetv3_model_fino.keras", "model_type": "mobilenetv3", "label": "MobileNetV3"},
    ])
    # Keep priority loaded for generic usages
    _load_first_available_model("lesiones_priority", [
        {"filename": "efficientnetB0.keras", "model_type": "efficientnet", "label": "EfficientNetB0"},
        {"filename": "efficientnet_model.keras", "model_type": "efficientnet", "label": "EfficientNet"},
        {"filename": "efficientnetB0.h5", "model_type": "efficientnet", "label": "EfficientNetB0"},
        {"filename": "resnet50_model_fine.keras", "model_type": "resnet50", "label": "ResNet50 Fine"},
        {"filename": "resnet50_model_fine.h5", "model_type": "resnet50", "label": "ResNet50"},
    ])
    yield

app = FastAPI(title=settings.app_name, version="0.3.0", lifespan=lifespan)

IMAGES_DIR.mkdir(parents=True, exist_ok=True)
app.mount("/images", StaticFiles(directory=str(IMAGES_DIR)), name="images")

# CORS Configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], # En prod, especificar dominios
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.post("/token")
async def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends()):
    try:
        async with httpx.AsyncClient() as client:
            print("ROBLE URL:", f"{settings.roble_auth_base}/{settings.roble_db_name}/login")
            print("EMAIL:", form_data.username)
            response = await client.post(
                
                f"{settings.roble_auth_base}/{settings.roble_db_name}/login",
                json={
                    "email": form_data.username,
                    "password": form_data.password,
                },
                timeout=10.0,
            )
        print("ROBLE STATUS:", response.status_code)
        print("ROBLE BODY:", response.text)
        if response.status_code not in (200, 201):
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Correo o contraseña incorrectos",
                headers={"WWW-Authenticate": "Bearer"},
            )

        data = response.json()
        
        access_token = data["accessToken"]

        current_user = User(
            username=data["user"]["email"],
            email=data["user"]["email"],
            role=data["user"]["role"],
            roble_user_id=data["user"]["id"],
        )

        #await ensure_user_exists(access_token, current_user)
        
        return {
            "access_token": data["accessToken"],
            "refresh_token": data["refreshToken"],
            "token_type": "bearer",
        }

    except httpx.RequestError:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="No se pudo conectar con ROBLE. Intenta nuevamente.",
        )


def _read_image_from_upload(file: UploadFile) -> np.ndarray:
    import cv2
    contents = file.file.read()
    arr = np.frombuffer(contents, dtype=np.uint8)
    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    if img is None:
        raise ValueError("El archivo no es una imagen válida o no pudo decodificarse.")
    return img


def _rd_risk_level(dr_grade: int) -> str:
    if dr_grade >= 3:
        return "high"
    if dr_grade > 0:
        return "medium"
    return "low"


def _encode_image_to_bytes(img_rgb: np.ndarray) -> bytes:
    """Convierte un array numpy RGB a bytes JPEG para guardado y modelos."""
    # OpenCV imencode espera BGR
    img_bgr = cv2.cvtColor(img_rgb, cv2.COLOR_RGB2BGR)
    success, buffer = cv2.imencode('.jpg', img_bgr, [int(cv2.IMWRITE_JPEG_QUALITY), 95])
    if not success:
        raise ValueError("No se pudo codificar la imagen preprocesada.")
    return buffer.tobytes()


def _rd_recommendation_short(dr_grade: int) -> str:
    if dr_grade >= 4:
        return "Evaluación oftalmológica urgente."
    if dr_grade >= 3:
        return "Derivación prioritaria a oftalmología."
    if dr_grade > 0:
        return "Seguimiento oftalmológico recomendado."
    return "Control oftalmológico periódico."


def _predict_rd_fallback(image: np.ndarray, model_label: str) -> dict:
    lesions = detect_hemorrhages(image)
    lesion_count = len(lesions)
    mean_intensity = float(np.mean(image))
    std_intensity = float(np.std(image))

    score = (0.7 * lesion_count) + (std_intensity / 85.0) + ((255.0 - mean_intensity) / 255.0)
    if score < 1.0:
        predicted_class = 0
    elif score < 1.5:
        predicted_class = 1
    elif score < 2.0:
        predicted_class = 2
    elif score < 2.5:
        predicted_class = 3
    else:
        predicted_class = 4

    confidence_percent = max(55.0, min(92.0, 58.0 + score * 12.0))
    probs = [8.0, 8.0, 8.0, 8.0, 8.0]
    probs[predicted_class] = confidence_percent
    rest = (100.0 - confidence_percent) / 4.0
    for i in range(5):
        if i != predicted_class:
            probs[i] = round(rest, 2)

    return {
        "model_used": model_label,
        "predicted_class": predicted_class,
        "confidence_percent": round(confidence_percent, 2),
        "diagnosis": "Normal" if predicted_class == 0 else f"Retinopatía (Grado {predicted_class})",
        "clinical_description": f"Resultado estimado por fallback para {model_label}.",
        "raw_probabilities": [round(float(p), 2) for p in probs],
    }


def _normalize_rd_prediction(model_id: str, model_spec: dict, prediction: dict, model_loaded: bool, inference_time_ms: float) -> dict:
    dr_grade = int(prediction.get("predicted_class", 0))
    confidence = float(prediction.get("confidence_percent", 0.0))
    risk_level = _rd_risk_level(dr_grade)
    runtime_meta = MODEL_RUNTIME_METADATA.get(model_spec["model_key"], {})
    model_name = runtime_meta.get("label", model_spec["label"])
    return {
        "model_id": model_id,
        "model_name": model_name,
        "model_used": model_name,
        "model_loaded": model_loaded,
        "predicted_class": dr_grade,
        "confidence_percent": round(confidence, 2),
        "diagnosis": prediction.get("diagnosis", "Normal"),
        "clinical_description": prediction.get("clinical_description", ""),
        "raw_probabilities": prediction.get("raw_probabilities", [0, 0, 0, 0, 0]),
        "risk_level": risk_level,
        "recommendation_short": _rd_recommendation_short(dr_grade),
        "inference_time_ms": round(inference_time_ms, 2),
    }


def _build_rd_comparison_result(filename: str, selected_models: List[str], model_results: List[dict], analysis_timestamp: str) -> dict:
    positive_models = [item for item in model_results if item["predicted_class"] > 0]
    grades = [item["predicted_class"] for item in model_results]
    consensus_grade = max(set(grades), key=lambda grade: (grades.count(grade), grade)) if grades else 0
    risk_level = _rd_risk_level(consensus_grade)
    headline = (
        f"{len(positive_models)}/{len(model_results)} modelos detectan retinopatía diabética"
        if positive_models
        else "Los modelos seleccionados no detectan retinopatía diabética"
    )
    primary_result = max(model_results, key=lambda item: item["confidence_percent"]) if model_results else {
        "model_name": "N/A",
        "predicted_class": 0,
        "confidence_percent": 0.0,
        "diagnosis": "Sin resultado",
        "clinical_description": "",
        "raw_probabilities": [0, 0, 0, 0, 0],
        "model_loaded": False,
        "risk_level": "low",
        "recommendation_short": _rd_recommendation_short(0),
        "inference_time_ms": 0.0,
    }

    return {
        "filename": filename,
        "timestamp": analysis_timestamp,
        "selected_models": selected_models,
        "model_comparisons": model_results,
        "primary_result": primary_result,
        "comparison_summary": {
            "headline": headline,
            "positive_models": len(positive_models),
            "total_models": len(model_results),
            "consensus_grade": consensus_grade,
            "risk_level": risk_level,
            "recommendation_short": _rd_recommendation_short(consensus_grade),
        },
        "model_used": primary_result["model_name"],
        "model_loaded": primary_result["model_loaded"],
        "predicted_class": primary_result["predicted_class"],
        "confidence_percent": primary_result["confidence_percent"],
        "diagnosis": primary_result["diagnosis"],
        "clinical_description": primary_result["clinical_description"],
        "raw_probabilities": primary_result["raw_probabilities"],
        "recommendation": headline,
        "recommendation_short": _rd_recommendation_short(consensus_grade),
        "risk_level": risk_level,
        "explanation": {
            "dr_grade": consensus_grade,
            "dr_diagnosis": headline,
            "recommendation_short": _rd_recommendation_short(consensus_grade),
            "risk_level": risk_level,
        },
        "disclaimer": DISCLAIMER,
    }


def _cdr_interpretation(cdr: float) -> str:
    if cdr > 0.6:
        return "CDR elevado; considerar evaluación de glaucoma."
    if cdr > 0.5:
        return "CDR en rango límite; seguimiento recomendado."
    return "CDR dentro de rango habitual."


def _build_full_result(
    results_by_model: dict,
    models_used: List[str],
    inference_times_ms: dict,
    img: np.ndarray,
) -> dict:
    """Construye respuesta completa: recomendación, explicación, postprocesamiento, gráficas."""
    model_a = results_by_model.get("A")
    model_b = results_by_model.get("B")
    model_c = results_by_model.get("C")

    # Recomendación y explicación (compatibles con 1, 2 o 3 modelos)
    cdr = model_a["cdr"] if model_a else 0.0
    prob = model_b if model_b is not None else 0.0
    
    # model_c is now a classification dict from ml_manager (or None)
    dr_class = model_c["predicted_class"] if model_c else 0
    dr_conf = model_c["confidence_percent"] if model_c else 0.0
    dr_diag = model_c["diagnosis"] if model_c else "Normal"
    
    risk_level = "high" if (prob >= 0.6 or cdr > 0.6 or dr_class >= 3) else "medium" if (prob >= 0.4 or cdr > 0.5 or dr_class > 0) else "low"
    recommendation_short = "Evaluación oftalmológica urgente." if risk_level == "high" else "Evaluación recomendada." if risk_level != "low" else "Seguimiento habitual."

    recommendation = f"CDR: {cdr:.2f}. {_cdr_interpretation(cdr)} Probabilidad glaucoma: {prob * 100:.1f}%."
    if model_c and dr_class > 0:
        recommendation += f" Retinopatía Diabética (Grado {dr_class} - {dr_conf:.1f}%)."
    elif "C" in models_used:
        recommendation += " Sin signos de retinopatía detectados."

    explanation = {
        "cdr_interpretation": _cdr_interpretation(cdr),
        "glaucoma_risk_level": risk_level,
        "glaucoma_probability_percent": round(prob * 100, 1),
        "dr_grade": dr_class,
        "dr_diagnosis": dr_diag,
        "recommendation_short": recommendation_short,
    }

    # Postprocesamiento: reporte con etiquetas y probabilidades
    report = build_report(
        model_a_result=model_a,
        model_b_result=model_b,
        model_c_result=model_c,
        models_used=models_used,
    )
    probabilities = report.get("probabilities", {})
    if "glaucoma" not in probabilities and "B" in models_used:
        probabilities["glaucoma"] = prob
    graph_data = graph_data_for_frontend(probabilities, inference_times_ms)

    lesions_found = []
    # If using the ML manager classification, we won't have bounding boxes, 
    # but we can return the overall classification result for the UI table
    if model_c:
        lesions_found.append({
            "x_min": 0, "y_min": 0, "x_max": 0, "y_max": 0,
            "label": f"Grado {model_c['predicted_class']}: {model_c['diagnosis']}", 
            "confidence": model_c['confidence_percent']/100.0,
        })

    return {
        "glaucoma_probability": round(prob, 4),
        "cup_to_disc_ratio": cdr,
        "disc_area": model_a.get("disc_area") if model_a else None,
        "cup_area": model_a.get("cup_area") if model_a else None,
        "lesions_found": lesions_found,
        "recommendation": recommendation,
        "explanation": explanation,
        "postprocessing": {
            "report": report,
            "graph_data": graph_data,
        },
        "disclaimer": DISCLAIMER,
    }


@app.get("/")
async def root():
    return {"service": settings.app_name, "docs": "/docs", "disclaimer": DISCLAIMER}


@app.get("/health")
async def health():
    return {"status": "ok", "service": settings.app_name, "version": "0.3.0"}


@app.get("/stats")
async def stats(
    current_user: User = Depends(get_current_user),
    token: str = Depends(oauth2_scheme),
):
    try:
        return await get_user_stats_from_roble(
            token=token,
            user_email=current_user.email,
        )
    except Exception as e:
        print("ERROR CARGANDO STATS DESDE ROBLE:", str(e))
        return get_global_stats()


@app.get("/history")
async def history(
    limit: int = Query(50, ge=1, le=100),
    offset: int = Query(0, ge=0),
    current_user: User = Depends(get_current_user),
    token: str = Depends(oauth2_scheme),
):
    try:
        inferences = await list_user_analyses_from_roble(
            token=token,
            user_email=current_user.email,
            limit=limit,
            offset=offset,
        )
        
        print("PRIMERA INFERENCIA HISTORY:", inferences[0] if inferences else "VACIO")
        return {"inferences": inferences}
        

    except Exception as e:
        print("ERROR LEYENDO HISTORIAL DESDE ROBLE:", str(e))
        return {
            "inferences": list_inferences(
                limit=limit,
                offset=offset,
                user_email=current_user.email,
            ),
            "source": "local_fallback",
        }


@app.delete("/history")
async def reset_history(
    current_user: User = Depends(get_current_user),
    token: str = Depends(oauth2_scheme),
):
    """Elimina historial completo del usuario autenticado."""
    try:
        result = await roble_delete_all_user_history(
            token=token,
            user_email=current_user.email,
        )

        return {
            "ok": True,
            "message": "Historial reiniciado correctamente.",
            **result,
        }

    except Exception as e:
        print("ERROR ELIMINANDO HISTORIAL:", str(e))

        raise HTTPException(
            status_code=500,
            detail="No se pudo limpiar el historial.",
        )


@app.delete("/history/{inference_id}")
async def delete_single_inference(
    inference_id: str,
    current_user: User = Depends(get_current_user),
    token: str = Depends(oauth2_scheme),
):
    try:
        result = await roble_delete_record(
            token=token,
            table_name="analisis_retina",
            id_column="inference_id",
            id_value=inference_id,
        )

        return {
            "ok": True,
            "message": "Análisis eliminado correctamente.",
            "deleted": result,
        }

    except Exception as e:
        print("ERROR ELIMINANDO ANALISIS EN ROBLE:", str(e))
        raise HTTPException(
            status_code=500,
            detail="No se pudo eliminar el análisis en la base de datos.",
        )


@app.delete("/batches/{batch_id}")
async def delete_entire_batch(
    batch_id: str,
    current_user: User = Depends(get_current_user),
    token: str = Depends(oauth2_scheme),
):
    """Elimina todas las inferencias asociadas a un lote en ROBLE DB."""
    try:
        result = await roble_delete_batch(
            token=token,
            batch_id=batch_id,
            user_email=current_user.email,
        )

        if result["deleted_count"] == 0:
            raise HTTPException(status_code=404, detail="Lote no encontrado o vacío")

        return {
            "ok": True,
            "message": f"Lote eliminado correctamente ({result['deleted_count']} análisis).",
            "deleted": result,
        }

    except HTTPException:
        raise

    except Exception as e:
        print("ERROR ELIMINANDO LOTE EN ROBLE:", str(e))
        raise HTTPException(
            status_code=500,
            detail="No se pudo eliminar el lote en la base de datos.",
        )


@app.get("/batches/{batch_id}")
async def get_batch_by_id(
    batch_id: str,
    current_user: User = Depends(get_current_user),
    token: str = Depends(oauth2_scheme),
):
    """Obtiene todas las inferencias de un lote desde ROBLE DB."""
    records = await get_batch_from_roble(
        token=token,
        batch_id=batch_id,
        user_email=current_user.email,
    )

    if not records:
        return JSONResponse(status_code=404, content={"detail": "Batch not found"})

    return records

@app.get("/export/batch/{batch_id}/excel")
async def export_batch_to_excel(batch_id: str, current_user: User = Depends(get_current_user)):
    """Exporta los datos de un lote a un archivo Excel."""
    records = get_batch(batch_id)
    if not records:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    # Procesar los datos según la lógica proporcionada
    json_main = {}
    all_models = set()
    for record in records:
        result = record.get("result", {})
        filename = result.get("filename", "unknown")
        model_comparisons = result.get("model_comparisons", [])
        
        dic_per_infe = {}
        for comparison in model_comparisons:
            comparison_model_name = comparison.get("model_id", "unknown")
            comparison_probs = comparison.get("raw_probabilities", [])
            dic_per_infe[comparison_model_name] = comparison_probs
            all_models.add(comparison_model_name)
        
        json_main[filename] = dic_per_infe
    
    models = sorted(list(all_models))
    
    # Obtener fecha del batch
    if records:
        batch_timestamp = records[0]["timestamp"]
        batch_date = datetime.fromisoformat(batch_timestamp).strftime('%Y-%m-%d %H:%M:%S')
    else:
        batch_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Etiquetas para grados
    grade_labels = ["NO R.D", "Leve", "Moderado", "Severo", "Proliferativo"]
    
    # Crear el Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Batch Data"
    
    num_columns = len(models) + 1  # Nombre + models
    last_col = openpyxl.utils.get_column_letter(num_columns)
    
    # Info del batch
    ws.merge_cells(f'A1:{last_col}1')
    cell = ws['A1']
    cell.value = "OCULAR-AI"
    cell.font = Font(size=24, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(f'A2:{last_col}2')
    cell = ws['A2']
    cell.value = f"Lote de {len(json_main)} archivos - {batch_date}"
    cell.font = Font(size=12, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.append([])
    
    # Encabezados
    headers = ["Nombre"] + models
    ws.append(headers)
    
    # Aplicar color azul a los encabezados
    fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    for col_num in range(1, num_columns + 1):
        cell = ws.cell(row=4, column=col_num)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    for filename, model_data in json_main.items():
        row = [filename]
        for model in models:
            probs = model_data.get(model, [0]*5)
            # Formatear como texto con etiquetas y porcentajes
            formatted = "\n".join([
                f"{label}: {prob}%" for label, prob in zip(grade_labels, probs[:5])
            ])
            row.append(formatted)
        ws.append(row)
    
    # Centrar todas las celdas y ajustar formato
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Ajustar ancho de columnas
    for col_num in range(1, num_columns + 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 25
    
    # Guardar en BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=batch_{batch_id}.xlsx"}
    )


@app.get("/export-pdf/{inference_id}")
async def export_pdf(inference_id: str, current_user: User = Depends(get_current_user)):
    from fpdf import FPDF
    import base64
    from io import BytesIO

    record = get_inference(inference_id)
    if not record:
        raise HTTPException(status_code=404, detail="Inference not found")

    result = record.get("result", {})
    summary = result.get("comparison_summary", {})
    primary = result.get("primary_result", {})
    timestamp = record.get("timestamp")
    
    class PDF(FPDF):
        def header(self):
            self.set_fill_color(0, 102, 204)
            self.rect(0, 0, 210, 40, 'F')
            self.set_font('Arial', 'B', 24)
            self.set_text_color(255, 255, 255)
            self.cell(0, 20, 'OCULAR-AI REPORT', 0, 1, 'C')
            self.set_font('Arial', '', 10)
            self.cell(0, 0, 'Plataforma de Analisis de Retinografias', 0, 1, 'C')
            self.ln(15)

        def footer(self):
            self.set_y(-20)
            self.set_font('Arial', 'I', 8)
            self.set_text_color(128, 128, 128)
            self.cell(0, 10, f'Pagina {self.page_no()} | OcularAI Research Group - Report ID: {inference_id}', 0, 0, 'C')

    pdf = PDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    # Seccion 1: Info General
    pdf.set_y(50)
    pdf.set_font('Arial', 'B', 14)
    pdf.set_text_color(0, 51, 102)
    pdf.cell(0, 10, 'RESUMEN DEL ANALISIS', 0, 1, 'L')
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(5)

    pdf.set_font('Arial', '', 10)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(40, 7, 'Archivo:', 0, 0)
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(0, 7, result.get("filename", "N/A"), 0, 1)
    
    pdf.set_font('Arial', '', 10)
    pdf.cell(40, 7, 'Fecha:', 0, 0)
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(0, 7, timestamp or "N/A", 0, 1)

    pdf.set_font('Arial', '', 10)
    pdf.cell(40, 7, 'ID Reporte:', 0, 0)
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(0, 7, inference_id, 0, 1)
    pdf.ln(5)

    # Seccion 2: Veredicto Clinico
    pdf.set_font('Arial', 'B', 14)
    pdf.set_text_color(0, 51, 102)
    pdf.cell(0, 10, 'VEREDICTO CLINICO', 0, 1, 'L')
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(5)

    risk_level = summary.get("risk_level") or result.get("risk_level", "low")
    risk_text = "ALTO" if risk_level == "high" else ("MEDIO" if risk_level == "medium" else "BAJO")
    
    pdf.set_font('Arial', 'B', 12)
    pdf.set_fill_color(240, 248, 255)
    pdf.cell(0, 15, f' RIESGO ESTIMADO: {risk_text}', 1, 1, 'L', fill=True)
    pdf.ln(5)

    pdf.set_font('Arial', 'B', 11)
    pdf.cell(0, 7, 'Diagnostico sugerido:', 0, 1)
    pdf.set_font('Arial', '', 11)
    pdf.multi_cell(0, 7, summary.get("headline") or primary.get("diagnosis") or "No se detecto patologia significativa.")
    pdf.ln(5)

    pdf.set_font('Arial', 'B', 11)
    pdf.cell(0, 7, 'Sugerencia medica:', 0, 1)
    pdf.set_font('Arial', 'I', 11)
    pdf.multi_cell(0, 7, f'"{summary.get("recommendation_short") or primary.get("recommendation_short") or "Se recomienda control periodico."}"')
    pdf.ln(10)

    # Seccion 3: Imagen (si existe)
    img_data = result.get("uploaded_image_preview")
    if img_data and "," in img_data:
        try:
            header, encoded = img_data.split(",", 1)
            img_bytes = base64.b64decode(encoded)
            img_stream = BytesIO(img_bytes)
            
            pdf.set_font('Arial', 'B', 14)
            pdf.set_text_color(0, 51, 102)
            pdf.cell(0, 10, 'RETINOGRAFIA ANALIZADA', 0, 1, 'L')
            pdf.line(10, pdf.get_y(), 200, pdf.get_y())
            pdf.ln(5)
            
            # Guardar temporalmente para ReportLab o usar BytesIO con FPDF2
            pdf.image(img_stream, x=45, w=120)
            pdf.ln(5)
        except Exception as e:
            pdf.cell(0, 10, f'No se pudo cargar la imagen: {str(e)}', 0, 1)

    # Seccion 4: Modelos
    pdf.add_page()
    pdf.set_font('Arial', 'B', 14)
    pdf.set_text_color(0, 51, 102)
    pdf.cell(0, 10, 'DETALLE POR MODELO IA', 0, 1, 'L')
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(5)

    comparisons = result.get("model_comparisons", [])
    if not comparisons:
        # Single result
        comparisons = [{
            "model_name": primary.get("model_used", "Modelo IA"),
            "diagnosis": primary.get("diagnosis", "N/A"),
            "confidence_percent": primary.get("confidence_percent", 0),
            "inference_time_ms": primary.get("inference_time_ms", 0)
        }]

    for comp in comparisons:
        pdf.set_font('Arial', 'B', 11)
        pdf.set_fill_color(245, 245, 245)
        pdf.cell(0, 8, f' {comp.get("model_name")}', 0, 1, 'L', fill=True)
        pdf.set_font('Arial', '', 10)
        pdf.cell(40, 7, 'Diagnostico:', 0, 0)
        pdf.cell(0, 7, comp.get("diagnosis", "N/A"), 0, 1)
        pdf.cell(40, 7, 'Confianza:', 0, 0)
        pdf.cell(0, 7, f'{float(comp.get("confidence_percent", 0)):.1f}%', 0, 1)
        pdf.cell(40, 7, 'Latencia:', 0, 0)
        pdf.cell(0, 7, f'{float(comp.get("inference_time_ms", 0)):.2f} ms', 0, 1)
        pdf.ln(5)

    pdf_output = pdf.output(dest='S')
    return StreamingResponse(
        BytesIO(pdf_output),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=reporte_{inference_id}.pdf"}
    )


@app.get("/inferences/{inference_id}")
async def get_inference_by_id(
    inference_id: str,
    current_user: User = Depends(get_current_user),
    token: str = Depends(oauth2_scheme),
):
    """Obtiene el registro completo de una inferencia desde ROBLE DB."""
    record = await get_analysis_from_roble(
        token=token,
        inference_id=inference_id,
        user_email=current_user.email,
    )

    if not record:
        return JSONResponse(status_code=404, content={"detail": "Inference not found"})

    return record


@app.post("/analyze-densenet/")
async def analyze_densenet(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """
    Análisis rápido DEMO usando solo el modelo DenseNet169 para Retinopatía Diabética.
    
    - **file**: imagen retinográfica.
    - Devuelve resultados detallados del modelo DenseNet con grado, probabilidades y descripción clínica.
    """
    loop = asyncio.get_event_loop()
    
    try:
        # 1. Leer imagen
        img = _read_image_from_upload(file)
        
        # 2. Preprocesar (Aplica Recorte + Ben Graham + Resize 224x224)
        preprocessed_rgb = preprocess_fundus(img)
        
        # 3. Convertir a bytes para guardado e inferencia
        # Usamos la imagen preprocesada para todo de ahora en adelante
        image_bytes = _encode_image_to_bytes(preprocessed_rgb)
        image_id = save_image_to_disk(image_bytes, file.filename)

        # 3. Ejecutar modelo DenseNet
        t0 = time.perf_counter()
        
        model_loaded = "lesiones_densenet" in ml_manager.models

        if model_loaded:
            # Ejecutar predicción real
            model_result = await loop.run_in_executor(
                None,
                lambda: ml_manager.predict_single(
                    model_key="lesiones_densenet",
                    image_bytes=image_bytes,
                    target_size=None,
                    model_type="densenet169"
                )
            )
        else:
            # Fallback para demo
            file.file.seek(0)
            image_bytes = file.file.read()
            image_id = save_image_to_disk(image_bytes, file.filename)
            
            lesions = await loop.run_in_executor(None, lambda: detect_hemorrhages(image))
            lesion_count = len(lesions)
            mean_intensity = float(np.mean(image))
            std_intensity = float(np.std(image))

            # Heurística simple para producir salida variable según imagen (0..4).
            score = (0.7 * lesion_count) + (std_intensity / 85.0) + ((255.0 - mean_intensity) / 255.0)
            if score < 1.0:
                predicted_class = 0
            elif score < 1.5:
                predicted_class = 1
            elif score < 2.0:
                predicted_class = 2
            elif score < 2.5:
                predicted_class = 3
            else:
                predicted_class = 4

            confidence_percent = max(55.0, min(92.0, 58.0 + score * 12.0))
            probs = [8.0, 8.0, 8.0, 8.0, 8.0]
            probs[predicted_class] = confidence_percent
            rest = (100.0 - confidence_percent) / 4.0
            for i in range(5):
                if i != predicted_class:
                    probs[i] = round(rest, 2)

            model_result = {
                "model_used": "FallbackHeuristic",
                "predicted_class": predicted_class,
                "confidence_percent": round(confidence_percent, 2),
                "diagnosis": "Normal" if predicted_class == 0 else f"Retinopatia (Grado {predicted_class})",
                "clinical_description": "Resultado estimado por fallback de demo (modelo no cargado).",
                "raw_probabilities": [round(float(p), 2) for p in probs],
            }
        
        inference_time_ms = (time.perf_counter() - t0) * 1000
        analysis_timestamp = datetime.now(timezone.utc).isoformat()
        
        # 4. Construir respuesta formateada
        dr_class = model_result["predicted_class"]
        confidence = model_result["confidence_percent"]
        diagnosis = model_result["diagnosis"]
        clinical_desc = model_result["clinical_description"]
        probabilities = model_result["raw_probabilities"]
        
        # Determinar nivel de riesgo
        risk_level = "high" if dr_class >= 3 else "medium" if dr_class > 0 else "low"
        
        # Recomendación
        recommendation_short = "Evaluación oftalmológica urgente." if risk_level == "high" else "Evaluación recomendada." if risk_level != "low" else "Seguimiento habitual."
        source_label = "DenseNet169" if model_loaded else "Fallback demo"
        recommendation = f"{source_label} predice: {diagnosis} (Confianza: {confidence:.1f}%). {clinical_desc}"
        
        # 5. Guardar trazabilidad
        inference_id = save_inference(
            models_used=["C"],
            inference_times_ms={"C": inference_time_ms},
            result={
                "predicted_class": dr_class,
                "confidence_percent": confidence,
                "diagnosis": diagnosis,
                "clinical_description": clinical_desc,
                "raw_probabilities": probabilities,
                "recommendation": recommendation,
                "timestamp": analysis_timestamp,
                "model_used": source_label,
                "risk_level": risk_level,
                "recommendation_short": recommendation_short,
                "filename": file.filename,
                "uploaded_image_preview": f"/images/{image_id}",
                "explanation": {
                    "dr_grade": dr_class,
                    "dr_diagnosis": diagnosis,
                    "recommendation_short": recommendation_short,
                    "glaucoma_risk_level": risk_level,
                }
            },
            image_size=(img.shape[1], img.shape[0]),
        )
        
        return {
            "success": True,
            "filename": file.filename,
            "uploaded_image_preview": f"/images/{image_id}",
            "inference_id": inference_id,
            "model_used": "DenseNet169" if model_loaded else "FallbackHeuristic",
            "predicted_class": dr_class,
            "confidence_percent": confidence,
            "diagnosis": diagnosis,
            "clinical_description": clinical_desc,
            "raw_probabilities": probabilities,
            "recommendation": recommendation,
            "recommendation_short": recommendation_short,
            "risk_level": risk_level,
            "timestamp": analysis_timestamp,
            "inference_time_ms": round(inference_time_ms, 2),
            "traceability": {
                "inference_id": inference_id,
                "models_used": ["C"],
                "inference_times_ms": {"C": round(inference_time_ms, 2)},
                "timestamp": analysis_timestamp,
            },
            "explanation": {
                "dr_grade": dr_class,
                "dr_diagnosis": diagnosis,
                "recommendation_short": recommendation_short,
                "glaucoma_risk_level": risk_level,
                "glaucoma_probability_percent": 0.0,  # No usamos modelo B
            },
            "model_loaded": model_loaded,
            "disclaimer": DISCLAIMER,
        }
    
    except Exception as e:
        return JSONResponse(
            status_code=400,
            content={"detail": f"Error procesando imagen: {str(e)}"}
        )


@app.post("/analyze-rd-comparison/")
async def analyze_rd_comparison(
    request: Request,
    files: List[UploadFile] = File(...),
    models: str = Query(
        "densenet169,resnet50,xception,efficientnet,mobilenetv3",
        description="Modelos RD a ejecutar.",
    ),
    current_user: User = Depends(get_current_user),
    token: str = Depends(oauth2_scheme),
):
    """
    Analiza una o más retinografías comparando 1, 2 o 3 modelos de retinopatía diabética.
    """
    try:
        return await _analyze_rd_comparison_impl(request, files, models, current_user, token)
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"[analyze-rd-comparison] Error: {e}\n{tb}", flush=True)
        return JSONResponse(
            status_code=500,
            content={"detail": f"Error en el servidor al procesar el lote: {str(e)}"},
        )


async def _analyze_rd_comparison_impl(
    request: Request,
    files: List[UploadFile],
    models: str,
    current_user: User,
    token: str,
):
    selected_models = [model.strip().lower() for model in models.split(",") if model.strip()]
    if not selected_models:
        selected_models = list(RD_MODEL_SPECS.keys())

    for model_id in selected_models:
        if model_id not in RD_MODEL_SPECS:
            return JSONResponse(
                status_code=400,
                content={"detail": f"Modelo inválido: {model_id}. Use densenet169, resnet50, xception, efficientnet y/o mobilenetv3."},
            )

    loop = asyncio.get_running_loop()
    final_results = []
    inferences_to_save = []
    batch_id = str(uuid.uuid4())
    cancelled = False
    # No limpiar imágenes para que no desaparezca el historial al recargar
    # clear_images_dir()

    for file in files:
        if await request.is_disconnected():
            cancelled = True
            break

        try:
            from backend.preprocessing.fundus import crop_img, apply_ben_graham_rgb
            img = _read_image_from_upload(file)
            # 1. Pasar de BGR a RGB
            img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            # 2. Cropping (Recorte de bordes negros)
            img_cropped = crop_img(img_rgb)

            # Guardar solo la imagen Ben Graham original, para mostrar en el frontend
            ben_graham_only_rgb = apply_ben_graham_rgb(img_cropped)
            ben_graham_only_bytes = _encode_image_to_bytes(ben_graham_only_rgb)
            ben_graham_only_image_id = save_image_to_disk2(
                ben_graham_only_bytes,
                f"ben_graham_{file.filename}"
            )
            
            
            #img_cropped = crop_img(img_rgb)
            
            # 3. Resize (Esta es la versión que mostramos al usuario)
            display_rgb = cv2.resize(img_cropped, (224, 224), interpolation=cv2.INTER_CUBIC)
            
            # 4. Ben Graham (Esta es la versión que el modelo IA necesita)
            preprocessed_rgb = apply_ben_graham_rgb(display_rgb)
            
            # Convertir a bytes para guardado e inferencia
            image_bytes = _encode_image_to_bytes(preprocessed_rgb)
            image_id = save_image_to_disk(image_bytes, file.filename)

            model_results = []
            inference_times_ms = {}
            analysis_timestamp = datetime.now(timezone.utc).isoformat()

            for model_id in selected_models:
                if await request.is_disconnected():
                    cancelled = True
                    break

                model_spec = RD_MODEL_SPECS[model_id]
                model_loaded = model_spec["model_key"] in ml_manager.models
                t0 = time.perf_counter()

                if model_loaded:
                    runtime_meta = MODEL_RUNTIME_METADATA.get(model_spec["model_key"], {})
                    prediction = await loop.run_in_executor(
                        None,
                        lambda model_spec=model_spec, runtime_meta=runtime_meta: ml_manager.predict_single(
                            model_key=model_spec["model_key"],
                            image_bytes=image_bytes,
                            target_size=None,
                            model_type=runtime_meta.get("model_type", model_spec["model_type"]),
                        ),
                    )
                else:
                    prediction = await loop.run_in_executor(
                        None,
                        lambda model_label=model_spec["label"]: _predict_rd_fallback(preprocessed_rgb, model_label),
                    )

                inference_time_ms = (time.perf_counter() - t0) * 1000
                inference_times_ms[model_id] = round(inference_time_ms, 2)
                model_results.append(
                    _normalize_rd_prediction(model_id, model_spec, prediction, model_loaded, inference_time_ms)
                )

            if cancelled or await request.is_disconnected():
                cancelled = True
                break

            result = _build_rd_comparison_result(file.filename, selected_models, model_results, analysis_timestamp)
            
            # Enviar solo las rutas relativas para evitar problemas de memoria en el frontend
            _, buffer = cv2.imencode(".jpg", img)
            original_img_id = save_image_to_disk(buffer.tobytes(), f"orig_{file.filename}")
            
            result["uploaded_image_preview"] = f"/images/{original_img_id}"
            result["uploaded_image_braham"] = f"/images/{ben_graham_only_image_id}"

            if await request.is_disconnected():
                cancelled = True
                break

            inference_id = str(uuid.uuid4())
            
            result["inference_id"] = inference_id
            result["batch_id"] = batch_id
            result["success"] = True
            result["traceability"] = {
                "inference_id": inference_id,
                "models_used": selected_models,
                "inference_times_ms": inference_times_ms,
                "timestamp": analysis_timestamp,
            }
            final_results.append(result)
            
            inferences_to_save.append({
                "inference_id": inference_id,
                "timestamp": analysis_timestamp,
                "models_used": selected_models,
                "inference_times_ms": inference_times_ms,
                "result": result,
                "image_size": (img.shape[1], img.shape[0]),
                "batch_id": batch_id,
                "user_email": current_user.email,
                "roble_user_id": current_user.roble_user_id,
            })
            
            roble_record = {
                "inference_id": inference_id,
                "roble_user_id": current_user.roble_user_id,
                "usuario_email": current_user.email,
                "batch_id": batch_id,
                "filename": file.filename,
                "timestamp": analysis_timestamp,
                "image_size": {
                    "width": img.shape[1],
                    "height": img.shape[0],
                },
                "risk_level": result.get("risk_level"),
                "diagnosis": result.get("diagnosis"),
                "confidence_percent": result.get("confidence_percent"),
                "recommendation_short": result.get("recommendation_short"),
                "result_json": result,
            }

            roble_response = await save_analysis_to_roble(
                token=token,
                record=roble_record,
            )

            print("RESPUESTA ROBLE INSERT:", roble_response)
            print("ANALISIS GUARDADO EN ROBLE OK")

        except Exception as e:
            final_results.append({
                "filename": file.filename,
                "error": f"Error interno procesando archivo: {str(e)}",
                "success": False,
            })

    if cancelled:
        return JSONResponse(
            status_code=499,
            content={"detail": "Analisis cancelado por el cliente."},
        )

    # Guardado masivo para optimizar I/O
    if inferences_to_save:
        save_inferences_batch(inferences_to_save)

    return final_results


@app.post("/analyze-demo/")
async def analyze_demo(
    file: UploadFile = File(...),
    model: str = Query("densenet169", description="Modelo a usar: densenet169, resnet50 o xception."),
):
    """
    Análisis demo sin autenticación. Solo 1 imagen, 1 modelo, sin guardar historial.
    """
    model_id = model.strip().lower()
    if model_id not in RD_MODEL_SPECS:
        return JSONResponse(
            status_code=400,
            content={"detail": f"Modelo inválido: {model_id}. Use densenet169, resnet50 o xception."},
        )

    loop = asyncio.get_event_loop()
    try:
        img = _read_image_from_upload(file)
        # Preprocesar
        preprocessed_rgb = preprocess_fundus(img)
        
        # Convertir a bytes para guardado e inferencia
        image_bytes = _encode_image_to_bytes(preprocessed_rgb)
        image_id = save_image_to_disk(image_bytes, file.filename)

        model_spec = RD_MODEL_SPECS[model_id]
        model_loaded = model_spec["model_key"] in ml_manager.models
        t0 = time.perf_counter()

        if model_loaded:
            runtime_meta = MODEL_RUNTIME_METADATA.get(model_spec["model_key"], {})
            prediction = await loop.run_in_executor(
                None,
                lambda: ml_manager.predict_single(
                    model_key=model_spec["model_key"],
                    image_bytes=image_bytes,
                    target_size=None,
                    model_type=runtime_meta.get("model_type", model_spec["model_type"]),
                ),
            )
        else:
            prediction = await loop.run_in_executor(
                None,
                lambda: _predict_rd_fallback(preprocessed_rgb, model_spec["label"]),
            )

        inference_time_ms = (time.perf_counter() - t0) * 1000
        analysis_timestamp = datetime.now(timezone.utc).isoformat()
        model_result = _normalize_rd_prediction(model_id, model_spec, prediction, model_loaded, inference_time_ms)

        return {
            "success": True,
            "filename": file.filename,
            "uploaded_image_preview": f"/images/{image_id}",
            "model_used": model_result["model_name"],
            "model_loaded": model_loaded,
            "predicted_class": model_result["predicted_class"],
            "confidence_percent": model_result["confidence_percent"],
            "diagnosis": model_result["diagnosis"],
            "clinical_description": model_result["clinical_description"],
            "raw_probabilities": model_result["raw_probabilities"],
            "risk_level": model_result["risk_level"],
            "recommendation_short": model_result["recommendation_short"],
            "inference_time_ms": round(inference_time_ms, 2),
            "timestamp": analysis_timestamp,
            "explanation": {
                "dr_grade": model_result["predicted_class"],
                "dr_diagnosis": model_result["diagnosis"],
                "recommendation_short": model_result["recommendation_short"],
                "risk_level": model_result["risk_level"],
            },
            "disclaimer": DISCLAIMER,
        }
    except Exception as e:
        return JSONResponse(
            status_code=400,
            content={"detail": f"Error procesando imagen: {str(e)}"},
        )


@app.post("/analyze-retina/")
async def analyze_retina(
    files: List[UploadFile] = File(...),
    models: str = Query(
        "A,B,C",
        description="Modelos a ejecutar: A (segmentación disco/copa), B (clasificación glaucoma), C (detección lesiones). Ej: A,B o B,C",
    ),
    model_c_type: str = Query(
        "resnet50v2",
        description="Tipo de modelo de IA a correr para las lesiones: 'resnet50v2' o 'mobilenetv3'"
    ),
    current_user: User = Depends(get_current_user),
):
    """
    Endpoint legado conservado por compatibilidad con flujos anteriores.

    Para la comparación actual de modelos de retinopatía diabética,
    se recomienda usar `/analyze-rd-comparison/`.
    """
    models_used = [m.strip().upper() for m in models.split(",") if m.strip()]
    if not models_used:
        models_used = ["A", "B", "C"]
    
    # Validar modelos una sola vez
    for m in models_used:
        if m not in ("A", "B", "C"):
            return JSONResponse(
                status_code=400,
                content={"detail": f"Modelo inválido: {m}. Use A, B y/o C."},
            )

    loop = asyncio.get_event_loop()
    final_results = []
    batch_id = str(uuid.uuid4())
    
    # 1. Pre-process and Batch Infer for Model C (Lesions/DR) if selected
    dr_batch_results = []
    c_inference_time_total = 0
    t0_batch = time.perf_counter()
    
    if model_c_type == "mobilenetv3":
        c_model_key = "lesiones_mobilenet"
    elif model_c_type == "densenet169":
        c_model_key = "lesiones_densenet"
    else:
        c_model_key = "lesiones_densenet"
    
    if "C" in models_used and c_model_key in ml_manager.models:
        valid_bytes_list = []
        for file in files:
            file.file.seek(0)
            valid_bytes_list.append(file.file.read())
            file.file.seek(0) # reset for cv2 read later
        
        # Run real model batch inference
        dr_batch_results = await loop.run_in_executor(
            None, 
            lambda: ml_manager.predict_batch(
                model_key=c_model_key, 
                images_bytes_list=valid_bytes_list, 
                model_type=model_c_type
            )
        )
        c_inference_time_total = (time.perf_counter() - t0_batch) * 1000

    for idx, file in enumerate(files):
        # Procesar cada imagen de forma independiente
        try:
            # 1. Leer imagen
            try:
                img = _read_image_from_upload(file)
            except Exception as e:
                final_results.append({
                    "filename": file.filename,
                    "error": str(e),
                    "success": False
                })
                continue

            # 2. Preprocesar (Aplica Recorte + Ben Graham + Resize 224x224)
            preprocessed_rgb = preprocess_fundus(img)
            
            # Convertir a bytes para guardado e inferencia
            image_bytes = _encode_image_to_bytes(preprocessed_rgb)
            image_id = save_image_to_disk(image_bytes, file.filename)
            
            # Usar la versión preprocesada (numpy) para los modelos legados A y B
            image = preprocessed_rgb

            # 3. Ejecutar modelos
            results_by_model = {}
            inference_times_ms = {}

            if "A" in models_used:
                t0 = time.perf_counter()
                seg = await loop.run_in_executor(None, lambda: segment_optic_disc(image))
                inference_times_ms["A"] = (time.perf_counter() - t0) * 1000
                results_by_model["A"] = seg

            if "B" in models_used:
                t0 = time.perf_counter()
                prob = await loop.run_in_executor(None, lambda: predict_glaucoma(image))
                inference_times_ms["B"] = (time.perf_counter() - t0) * 1000
                results_by_model["B"] = float(prob)

            if "C" in models_used:
                if dr_batch_results and len(dr_batch_results) > idx:
                    results_by_model["C"] = dr_batch_results[idx]
                    # distribute time evenly
                    inference_times_ms["C"] = c_inference_time_total / len(files)
                else:
                    # Fallback to dummy if real model was not loaded
                    t0 = time.perf_counter()
                    lesions = await loop.run_in_executor(None, lambda: detect_hemorrhages(image))
                    inference_times_ms["C"] = (time.perf_counter() - t0) * 1000
                    # format dummy detector output to match ml manager dict roughly
                    dummy_label = lesions[0]['label'] if lesions else 'Normal'
                    results_by_model["C"] = {
                        "predicted_class": 1 if lesions else 0,
                        "confidence_percent": lesions[0]['confidence']*100 if lesions else 100.0,
                        "diagnosis": dummy_label,
                        "clinical_description": "Mocked detector",
                        "raw_probabilities": [10.0]*5
                    }

            # Defaults
            if "A" not in results_by_model: results_by_model["A"] = None
            if "B" not in results_by_model: results_by_model["B"] = None
            if "C" not in results_by_model: results_by_model["C"] = None

            # 4. Construir resultado
            result = _build_full_result(
                results_by_model,
                models_used,
                inference_times_ms,
                img,
            )
            result["uploaded_image_preview"] = f"/images/{image_id}"

            # 5. Guardar trazabilidad
            inference_id = save_inference(
                models_used=models_used,
                inference_times_ms=inference_times_ms,
                result=result,
                image_size=(img.shape[1], img.shape[0]),
                batch_id=batch_id,
                user_email=current_user.email,
                roble_user_id=current_user.roble_user_id,
            )

            result["inference_id"] = inference_id
            result["filename"] = file.filename
            result["success"] = True
            result["traceability"] = {
                "inference_id": inference_id,
                "models_used": models_used,
                "inference_times_ms": inference_times_ms,
            }
            
            final_results.append(result)

        except Exception as e:
            final_results.append({
                "filename": file.filename,
                "error": f"Error interno procesando archivo: {str(e)}",
                "success": False
            })

    return final_results


# ---------------------------------------------------------------------------
# Agente Cerebro: orquestador LLM que decide dinámicamente qué modelos usar
# ---------------------------------------------------------------------------

@app.post("/analyze-agent/", response_model=AgentAnalysisResponse)
async def analyze_agent(
    file: UploadFile = File(...),
    analysis_request: str = Query(
        default=(
            "Realiza un análisis retiniano completo que cubra retinopatía diabética, "
            "riesgo de glaucoma y evaluación del disco óptico."
        ),
        description=(
            "Instrucción en lenguaje natural para el agente cerebro. "
            "El agente decidirá qué modelos ML invocar según la solicitud."
        ),
    ),
    current_user: User = Depends(get_current_user),
):
    """
    Endpoint del agente cerebro. Claude orquesta los modelos A/B/C dinámicamente
    según la solicitud de análisis. Los resultados se guardan en el store para trazabilidad.
    """
    try:
        image_bytes = await file.read()

        arr = np.frombuffer(image_bytes, dtype=np.uint8)
        import cv2 as _cv2
        img = _cv2.imdecode(arr, _cv2.IMREAD_COLOR)
        if img is None:
            return JSONResponse(
                status_code=400,
                content={"detail": "El archivo no es una imagen válida."},
            )
        image_size = (img.shape[1], img.shape[0])

        image_id = save_image_to_disk(image_bytes, file.filename)

        result = await brain_agent.run(
            image_bytes=image_bytes,
            filename=file.filename,
            image_id=image_id,
            image_size=image_size,
            analysis_request=analysis_request,
        )
        return result

    except Exception as e:
        if anthropic_sdk is not None and isinstance(e, anthropic_sdk.AuthenticationError):
            return JSONResponse(
                status_code=401,
                content={"detail": "ANTHROPIC_API_KEY inválida o no configurada."},
            )
        if anthropic_sdk is not None and isinstance(e, anthropic_sdk.RateLimitError):
            return JSONResponse(
                status_code=429,
                content={"detail": "Rate limit de la API de Anthropic. Intente nuevamente."},
            )
        if anthropic_sdk is not None and isinstance(e, anthropic_sdk.APIStatusError):
            return JSONResponse(
                status_code=502,
                content={"detail": f"Error de API Anthropic ({e.status_code}): {e.message}"},
            )
        if isinstance(e, ModuleNotFoundError) and getattr(e, "name", None) == "anthropic":
            return JSONResponse(
                status_code=503,
                content={
                    "detail": "Falta el paquete anthropic. En la raíz del repo: pip install -r requirements.txt"
                },
            )
        return JSONResponse(
            status_code=500,
            content={"detail": f"Error interno del agente cerebro: {str(e)}"},
        )

@app.post("/logout")
async def logout(token: str = Depends(oauth2_scheme)):
    try:
        async with httpx.AsyncClient() as client:
            response = await client.post(
                f"{settings.roble_auth_base}/{settings.roble_db_name}/logout",
                headers={
                    "Authorization": f"Bearer {token}"
                },
                timeout=10.0,
            )

        if response.status_code not in (200, 201):
            raise HTTPException(
                status_code=400,
                detail="No se pudo cerrar sesión en ROBLE"
            )

        return {"message": "Sesión cerrada correctamente"}

    except httpx.RequestError:
        raise HTTPException(
            status_code=503,
            detail="Error conectando con ROBLE"
        )
        
@app.post("/test-roble-db")
async def test_roble_db(
    token: str = Depends(oauth2_scheme),
    current_user: User = Depends(get_current_user),
):
    record = {
        "roble_user_id": current_user.roble_user_id,
        "email": current_user.email,
        "nombre": current_user.username,
        "rol": current_user.role,
        "activo_app": True,
        "fecha_creacion": datetime.now(timezone.utc).isoformat(),
        "ultimo_login": datetime.now(timezone.utc).isoformat(),
    }

    result = await roble_insert(
        token=token,
        table_name="usuarios_app",
        records=[record],
    )

    return {
        "ok": True,
        "insert_result": result,
        "record": record,
    }